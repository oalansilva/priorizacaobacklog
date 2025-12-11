"""Serviço de exportação de roadmaps para Excel."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.db import Roadmap
from datetime import datetime


def export_roadmap_to_excel(roadmap: Roadmap) -> BytesIO:
    """
    Gera arquivo Excel do roadmap com formatação profissional.
    
    Args:
        roadmap: Objeto Roadmap com dados da priorização
        
    Returns:
        BytesIO com o arquivo Excel gerado
    """
    wb = Workbook()
    
    # Cores
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    priorizado_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    despriorizado_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === ABA 1: RESUMO ===
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo['A1'] = "Roadmap - Resumo da Priorização"
    ws_resumo['A1'].font = title_font
    ws_resumo.merge_cells('A1:D1')
    
    # Data
    data_criacao = datetime.fromisoformat(roadmap.created_at).strftime("%d/%m/%Y %H:%M")
    ws_resumo['A2'] = f"Data: {data_criacao}"
    ws_resumo.merge_cells('A2:D2')
    
    # Métricas
    row = 4
    metricas = [
        ("Capacidade Total", f"{roadmap.capacidade_total}h"),
        ("Percentual Sustentação", f"{roadmap.percentual_sustentacao}%"),
        ("Capacidade Iniciativas", f"{roadmap.capacidade_iniciativas}h"),
        ("", ""),
        ("Total de Itens", roadmap.total_itens),
        ("Itens Priorizados", roadmap.itens_priorizados),
        ("Itens Despriorizados", roadmap.itens_despriorizados),
        ("Horas Alocadas", f"{roadmap.horas_alocadas}h"),
        ("", ""),
        ("Pesos Utilizados", ""),
        ("  Financeiro", f"{roadmap.peso_financeiro}%"),
        ("  Negócios", f"{roadmap.peso_negocios}%"),
        ("  Cliente", f"{roadmap.peso_cliente}%"),
        ("  OKR", f"{roadmap.peso_okr}%"),
    ]
    
    for label, value in metricas:
        ws_resumo[f'A{row}'] = label
        ws_resumo[f'B{row}'] = value
        if label and not label.startswith("  "):
            ws_resumo[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 20
    
    # === ABA 2: ITENS PRIORIZADOS ===
    ws_priorizados = wb.create_sheet("Priorizados")
    
    # Cabeçalho
    headers = ["#", "Título", "Área", "Esforço (h)", "Score", "Impacto Fin.", "Impacto Neg.", "Impacto Cli.", "OKR", "Must Have"]
    for col, header in enumerate(headers, 1):
        cell = ws_priorizados.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    priorizados = [item for item in roadmap.itens if item.status == "Priorizado"]
    priorizados.sort(key=lambda x: x.prioridade if x.prioridade else 999)
    
    for idx, item in enumerate(priorizados, 2):
        ws_priorizados.cell(idx, 1, item.prioridade or "").border = border
        ws_priorizados.cell(idx, 2, item.titulo or "").border = border
        ws_priorizados.cell(idx, 3, item.area or "").border = border
        ws_priorizados.cell(idx, 4, item.esforco_estimado or 0).border = border
        
        # Tratar score que pode ser None ou float
        score_value = item.score if item.score is not None else 0.0
        ws_priorizados.cell(idx, 5, f"{score_value:.1f}%").border = border
        
        ws_priorizados.cell(idx, 6, item.impacto_financeiro or "Não").border = border
        ws_priorizados.cell(idx, 7, item.impacto_negocios or "Não").border = border
        ws_priorizados.cell(idx, 8, item.impacto_cliente or "Não").border = border
        ws_priorizados.cell(idx, 9, item.okr or "Não").border = border
        ws_priorizados.cell(idx, 10, item.must_have or "Não").border = border
        
        # Aplicar cor de fundo
        for col in range(1, 11):
            ws_priorizados.cell(idx, col).fill = priorizado_fill
    
    # Ajustar larguras
    ws_priorizados.column_dimensions['A'].width = 5
    ws_priorizados.column_dimensions['B'].width = 50
    ws_priorizados.column_dimensions['C'].width = 20
    ws_priorizados.column_dimensions['D'].width = 12
    ws_priorizados.column_dimensions['E'].width = 10
    for col in ['F', 'G', 'H', 'I', 'J']:
        ws_priorizados.column_dimensions[col].width = 12
    
    # === ABA 3: ITENS DESPRIORIZADOS ===
    ws_despriorizados = wb.create_sheet("Despriorizados")
    
    # Cabeçalho
    for col, header in enumerate(headers, 1):
        cell = ws_despriorizados.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    despriorizados = [item for item in roadmap.itens if item.status == "Despriorizado"]
    despriorizados.sort(key=lambda x: x.prioridade if x.prioridade else 999)
    
    for idx, item in enumerate(despriorizados, 2):
        ws_despriorizados.cell(idx, 1, item.prioridade or "").border = border
        ws_despriorizados.cell(idx, 2, item.titulo or "").border = border
        ws_despriorizados.cell(idx, 3, item.area or "").border = border
        ws_despriorizados.cell(idx, 4, item.esforco_estimado or 0).border = border
        
        # Tratar score que pode ser None ou float
        score_value = item.score if item.score is not None else 0.0
        ws_despriorizados.cell(idx, 5, f"{score_value:.1f}%").border = border
        
        ws_despriorizados.cell(idx, 6, item.impacto_financeiro or "Não").border = border
        ws_despriorizados.cell(idx, 7, item.impacto_negocios or "Não").border = border
        ws_despriorizados.cell(idx, 8, item.impacto_cliente or "Não").border = border
        ws_despriorizados.cell(idx, 9, item.okr or "Não").border = border
        ws_despriorizados.cell(idx, 10, item.must_have or "Não").border = border
        
        # Aplicar cor de fundo
        for col in range(1, 11):
            ws_despriorizados.cell(idx, col).fill = despriorizado_fill
    
    # Ajustar larguras
    ws_despriorizados.column_dimensions['A'].width = 5
    ws_despriorizados.column_dimensions['B'].width = 50
    ws_despriorizados.column_dimensions['C'].width = 20
    ws_despriorizados.column_dimensions['D'].width = 12
    ws_despriorizados.column_dimensions['E'].width = 10
    for col in ['F', 'G', 'H', 'I', 'J']:
        ws_despriorizados.column_dimensions[col].width = 12
    
    # Salvar em BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
